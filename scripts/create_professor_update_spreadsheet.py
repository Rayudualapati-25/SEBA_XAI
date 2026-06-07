from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "spreadsheets"
OUT_FILE = OUT_DIR / "SEBA_XAI_Professor_Update_Details.xlsx"


SOURCE_FILES = [
    "professor_update_2026-05-27.md",
    "00_problem_understanding.md",
    "01_literature_review.md",
    "03_datasets.md",
    "05_research_gap.md",
    "06_proposed_architecture.md",
    "07_methodology.md",
    "08_experiment_plan.md",
    "09_evaluation_metrics.md",
    "10_ethics_security_legal.md",
    "13_implementation_kickstart.md",
    "14_existing_models_for_testing.md",
    "sources/source_log.md",
    "02_literature_matrix.csv",
    "04_dataset_matrix.csv",
]


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")
OK_FILL = PatternFill("solid", fgColor="E2F0D9")
LINK_FONT = Font(color="0563C1", underline="single")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BODY_FONT = Font(name="Calibri", size=11)
THIN = Side(style="thin", color="D9E2F3")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def link(url: str) -> str:
    return url


def add_sheet(wb: Workbook, title: str, headers: list[str], rows: Iterable[Iterable[object]]) -> None:
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(list(row))
    style_sheet(ws)


def style_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
            if isinstance(cell.value, str) and cell.value.startswith("http"):
                cell.hyperlink = cell.value
                cell.font = LINK_FONT
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 80))
        width = max(12, min(max_len + 2, 55))
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.sheet_view.showGridLines = False


def read_csv(path: Path) -> list[list[str]]:
    with path.open(newline="", encoding="utf-8") as fp:
        return list(csv.reader(fp))


def extract_markdown_lines(files: list[str]) -> list[list[object]]:
    rows = []
    for rel in files:
        path = ROOT / rel
        if not path.exists() or path.suffix.lower() != ".md":
            continue
        heading_stack: list[str] = []
        for line_no, line in enumerate(path.read_text(encoding="utf-8").splitlines(), start=1):
            stripped = line.strip()
            if stripped.startswith("#"):
                match = re.match(r"^(#+)\s+(.*)$", stripped)
                if match:
                    level = len(match.group(1))
                    title = match.group(2)
                    heading_stack = heading_stack[: level - 1]
                    heading_stack.append(title)
            if stripped:
                rows.append([rel, line_no, " > ".join(heading_stack), stripped])
    return rows


def extract_source_links() -> list[list[str]]:
    rows: list[list[str]] = []
    current_group = "General"
    for line in (ROOT / "sources/source_log.md").read_text(encoding="utf-8").splitlines():
        if line.startswith("## "):
            current_group = line.replace("#", "").strip()
        if line.strip().startswith("- ") and "http" in line:
            text = line.strip()[2:]
            parts = text.split(":", 1)
            label = parts[0].strip() if len(parts) > 1 else text
            urls = re.findall(r"https?://[^\s)]+", text)
            rows.append([current_group, label, text, urls[0] if urls else ""])
    return rows


def main() -> None:
    OUT_DIR.mkdir(exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    add_sheet(
        wb,
        "00_Read_Me",
        ["Field", "Detail"],
        [
            ["Workbook purpose", "Professor-facing spreadsheet for learning progress and SEBA-XAI research project status."],
            ["Created on", "27 May 2026"],
            ["Main research title", "SEBA-XAI: Secure Explainable Blockchain-Audited Access Overlay for Inter-Agency Police Data Sharing in India"],
            ["Core model direction", "Access-decision model: allow, deny, or escalate sensitive police/crime-record access requests."],
            ["Evidence boundary", "No implementation results, accuracy, latency, security-improvement, deployment, or legal-compliance claim is made yet."],
            ["Best next step", "Build synthetic_access_sim with baselines, audit logs, XAI explanation artifacts, tamper tests, and metrics."],
            ["How to read this workbook", "Start with Update_Map, then Learning tabs, Research_Work, Model_Workflow, Existing_Models, Literature_Matrix, Dataset_Matrix, Experiments, Metrics, and Ethics_Legal."],
            ["Traceability", "The Raw_MD_Lines sheet preserves lines from the supporting Markdown files."],
        ],
    )

    update_rows = [
        ["Learning Progress", "Blockchain", "Completed foundation topics and most consensus-algorithm concepts.", "Completed/partially complete", "professor_update_2026-05-27.md"],
        ["Learning Progress", "XAI", "Prepared XAI syllabus and notes; identified XAI role as access-decision explanation.", "Prepared; learning continues", "professor_update_2026-05-27.md"],
        ["Learning Progress", "Agentic AI", "Planned next learning block: agents, tool use, memory, RAG, multi-agent workflows, safety, audit logging.", "Pending/formal study next", "professor_update_2026-05-27.md"],
        ["Research Progress", "Current topic", "SEBA-XAI secure explainable blockchain-audited access overlay for inter-agency police data sharing in India.", "Selected direction", "professor_update_2026-05-27.md"],
        ["Research Progress", "CCTNS/ICJS boundary", "System supports existing systems; it does not replace CCTNS or ICJS.", "Important boundary", "professor_update_2026-05-27.md"],
        ["Research Progress", "Research question", "Should a sensitive record request be allowed, denied, or escalated, and can the decision be explained and audited later?", "Core question", "professor_update_2026-05-27.md"],
        ["Current Status", "Foundation complete", "Problem, literature, datasets, gap, architecture, methodology, experiments, metrics, ethics, and professor documents prepared.", "Completed research design", "professor_update_2026-05-27.md"],
        ["Current Status", "Implementation", "Implementation is not completed yet and no experimental result is claimed.", "Pending", "professor_update_2026-05-27.md"],
    ]
    add_sheet(wb, "01_Update_Map", ["Section", "Subsection", "Detail", "Status", "Source"], update_rows)

    blockchain_completed = [
        "hash functions", "Merkle trees", "digital signatures", "commitment schemes", "basic distributed systems",
        "CAP theorem", "FLP impossibility", "Bitcoin architecture", "blocks and transactions", "UTXO model",
        "Proof of Work basic idea", "consensus design space", "PBFT", "HotStuff", "Tendermint/CometBFT",
        "Ouroboros", "Algorand BA*", "Avalanche/Snowball", "Solana Proof of History", "Casper FFG basics",
        "slashing", "nothing-at-stake problem", "long-range attack basics",
    ]
    blockchain_remaining = [
        "Proof of Work variants: SHA-256, Ethash, RandomX, Scrypt, Equihash",
        "full comparison of PoW and PoS", "validator selection and validator rewards", "Proof of Authority",
        "consortium-chain consensus", "LMD-GHOST in Ethereum",
        "51% attack, selfish mining, eclipse attack, bribery attack, inactivity leak, time-bandit attack",
        "final consensus comparison table", "consensus lab and quiz",
    ]
    add_sheet(
        wb,
        "02_Blockchain",
        ["Status", "Topic", "Why it matters for SEBA-XAI"],
        [[ "Completed", t, "Supports audit ledger, cryptographic proof, consensus, and tamper-evidence understanding."] for t in blockchain_completed]
        + [[ "Remaining", t, "Needed for stronger consensus/security discussion and permissioned-chain comparison."] for t in blockchain_remaining],
    )

    xai_topics = [
        "what explainability means", "interpretable models and post-hoc explanations", "taxonomy of XAI methods",
        "LIME and SHAP", "gradient-based explanations", "concept-based explanations", "counterfactual explanations",
        "global explanation methods", "explanation evaluation", "human-centered XAI", "XAI in law and public safety",
    ]
    xai_research_uses = [
        "why access was allowed", "why access was denied", "why access was sent for superior approval",
        "which rule or attribute caused the decision", "whether the explanation can be checked later during audit",
    ]
    add_sheet(
        wb,
        "03_XAI",
        ["Category", "Detail", "Research connection"],
        [[ "Learning path", t, "Builds background for explanation artifacts and reviewability."] for t in xai_topics]
        + [[ "SEBA-XAI use", u, "Explains allow/deny/escalate decisions for sensitive access requests."] for u in xai_research_uses],
    )

    agentic_topics = [
        "what an AI agent is", "planning and task decomposition", "tool use", "memory",
        "retrieval-augmented generation", "multi-agent workflows", "evaluation of agent actions",
        "safety and guardrails", "audit logging of agent actions",
    ]
    add_sheet(
        wb,
        "04_Agentic_AI",
        ["Status", "Topic", "Research connection"],
        [["Planned next", t, "May support future officer/auditor/supervisor assistant; not the current core contribution."] for t in agentic_topics],
    )

    research_work = [
        ["Problem understanding", "Studied Indian policing context; identified CCTNS and ICJS; research supports existing systems instead of replacing them.", "Completed"],
        ["Literature review", "Reviewed blockchain evidence management, Hyperledger Fabric, RBAC/ABAC/PBAC, XAI/fairness in high-stakes AI, privacy-preserving ML, and secure sharing.", "Completed"],
        ["Existing model search", "Found exact full systems are rare; closest works include BAXDT, LEChain, and two-level blockchain evidence systems.", "Completed"],
        ["Dataset study", "Studied NCRB and BPRD; found NCRB is aggregate; selected synthetic multi-station access-request data for the core experiment.", "Completed"],
        ["Research gap", "Rejected raw police data on blockchain and unsupported individual crime prediction; selected access governance gap.", "Completed"],
        ["Architecture", "Designed off-chain raw records, on-chain hashes/audit info, RBAC/ABAC/PBAC, credential checks, superior approval, encrypted storage, and XAI explanations.", "Completed design"],
        ["Methodology and experiments", "Planned RBAC, ABAC/PBAC, signed hash-chain, and blockchain-audited SEBA-XAI baselines; planned tamper, latency, leakage, and explanation tests.", "Planned"],
        ["Ethics/legal boundary", "No deployment-ready, legal-compliance, sensitive police data, or operational-benefit claim is made.", "Boundary set"],
    ]
    add_sheet(wb, "05_Research_Work", ["Area", "Details", "Status"], research_work)

    model_rows = [
        ["Input", "officer role", "Subject attribute"],
        ["Input", "officer rank", "Subject attribute"],
        ["Input", "officer police station", "Subject attribute"],
        ["Input", "jurisdiction", "Subject/environment attribute"],
        ["Input", "case assignment", "Subject-object relation"],
        ["Input", "requested record type", "Object attribute"],
        ["Input", "sensitivity level", "Object attribute"],
        ["Input", "victim/witness/juvenile flag", "Sensitive object attribute"],
        ["Input", "purpose of access", "Environment/action attribute"],
        ["Input", "time window", "Environment attribute"],
        ["Input", "credential status", "Security attribute"],
        ["Input", "emergency flag", "Environment attribute"],
        ["Input", "approval-token status", "Approval/security attribute"],
        ["Output", "allow", "Access granted"],
        ["Output", "deny", "Access rejected"],
        ["Output", "escalate to superior approval", "Human review required"],
        ["Explanation", "which rule passed", "XAI/policy trace"],
        ["Explanation", "which rule failed", "XAI/policy trace"],
        ["Explanation", "why escalation was needed", "XAI/policy trace"],
        ["Explanation", "what approval is required", "Approval guidance"],
        ["Explanation", "which policy version was used", "Audit reconstruction"],
        ["Blockchain proof", "request hash", "On-chain audit commitment"],
        ["Blockchain proof", "policy hash", "On-chain audit commitment"],
        ["Blockchain proof", "decision hash", "On-chain audit commitment"],
        ["Blockchain proof", "approval hash", "On-chain audit commitment"],
        ["Blockchain proof", "model version", "Model audit metadata"],
        ["Blockchain proof", "explanation hash", "XAI audit commitment"],
        ["Blockchain proof", "timestamp", "Event ordering"],
        ["Blockchain proof", "actor credential hash", "Actor proof without raw credential exposure"],
    ]
    add_sheet(wb, "06_Model_Workflow", ["Type", "Field/Decision", "Meaning"], model_rows)

    existing_models = [
        [1, "Blockchain-assisted explainable decision traces (BAXDT)", 2025, "Blockchain + XAI decision traces", "SHAP explanations; explanation-density metric; blockchain anchoring; model/data context", "Very high for XAI artifact hashing and decision trace design", "Rebuild/adapt decision-trace idea locally", "Paper claims open source, repository not verified", link("https://doi.org/10.1016/j.knosys.2025.114402"), link("https://www.sciencedirect.com/science/article/pii/S0950705125014418"), ""],
        [2, "Blockchain-based auditing of legal decisions supported by XAI and generative AI tools", 2024, "Blockchain + XAI + legal decision audit", "Legal decision audit trail; XAI artifacts; Ethereum/Hyperledger Fabric evaluation pattern", "Very high for legal AI audit pattern", "Use as conceptual/evaluation precedent", "No public repo found; not police-record access control", link("https://doi.org/10.1016/j.engappai.2023.107666"), link("https://livrepository.liverpool.ac.uk/3179806/"), ""],
        [3, "LEChain", 2021, "Blockchain + lawful evidence management", "Consortium blockchain; CP-ABE; police-to-court evidence lifecycle; witness/juror privacy", "High for police/legal blockchain access workflow", "Inspect and adapt evidence/access-control ideas", "No XAI; not access-request justification", link("https://doi.org/10.1016/j.future.2020.09.038"), link("https://www.sciencedirect.com/science/article/pii/S0167739X1933167X"), link("https://github.com/SopmmmodII/LEChain")],
        [4, "Two-Level Blockchain System for Digital Crime Evidence Management", 2021, "Crime evidence blockchain architecture", "Police/prosecutor/court consortium; hot/cold blockchain split; Hyperledger Fabric", "High for evidence architecture", "Borrow hot/cold split and consortium design", "No XAI; no public code found", link("https://doi.org/10.3390/s21093051"), link("https://www.mdpi.com/1424-8220/21/9/3051"), ""],
        [5, "User authentication and access control to blockchain-based forensic log data", 2023, "Forensic log access control", "Authentication; access control; blockchain forensic logs; AVISPA verification", "Medium for security baseline", "Borrow RBAC/ABAC and protocol verification ideas", "No XAI; no public repo verified", link("https://doi.org/10.1186/s13635-023-00142-3"), link("https://link.springer.com/article/10.1186/s13635-023-00142-3"), ""],
        [6, "BlendSPS", 2020, "Blockchain-enabled smart public safety", "Public-safety microservices; blockchain security service; Ethereum/Tendermint prototype", "Medium for public-safety blockchain architecture", "Use for modular architecture inspiration", "Not XAI-centered; surveillance/IoT focus", link("https://doi.org/10.3390/smartcities3030047"), link("https://www.mdpi.com/2624-6511/3/3/47"), link("https://github.com/samuelxu999/Research/tree/master/Security/py_dev/BlendSPS/")],
        [7, "A Secure and Privacy-Preserving Blockchain-Based XAI-Justice System", 2023, "Blockchain + privacy + XAI justice framework", "Privacy techniques; explainable AI; NLP/legal decision support", "Medium for broad justice framing", "Cite as related vision", "Conceptual; not empirical baseline", link("https://doi.org/10.3390/info14090477"), link("https://www.mdpi.com/2078-2489/14/9/477"), ""],
        [8, "Blockchain for explainable and trustworthy artificial intelligence", "2019/2020", "General blockchain + XAI", "Smart contracts; trusted oracles; decentralized storage; AI/XAI consensus", "Medium as theory background", "Cite for general rationale only", "Not police/legal access governance implementation", link("https://doi.org/10.1002/widm.1340"), link("https://scholarworks.aub.edu.lb/handle/10938/25595?show=full"), ""],
    ]
    add_sheet(
        wb,
        "07_Existing_Models",
        ["Rank", "System/Paper", "Year", "Domain", "Components", "Fit", "Recommended Use", "Limitation", "DOI", "Source Link", "Code Link"],
        existing_models,
    )

    lit = read_csv(ROOT / "02_literature_matrix.csv")
    add_sheet(wb, "08_Literature_Matrix", lit[0], lit[1:])

    data = read_csv(ROOT / "04_dataset_matrix.csv")
    add_sheet(wb, "09_Dataset_Matrix", data[0], data[1:])

    architecture_rows = [
        ["Architecture principle", "Overlay, not replacement", "SEBA-XAI sits above/beside CCTNS/ICJS-style systems; raw records stay authoritative in agency systems."],
        ["Existing agency systems", "Source systems", "Police station/state CCTNS-like systems; courts, prisons, forensics, prosecution systems."],
        ["Access request gateway", "Request normalization", "Receives access requests and normalizes subject, object, action, and environment attributes."],
        ["Policy decision layer", "Security enforcement", "RBAC baseline; ABAC/PBAC policies; superior approval; revocation; time-window checks."],
        ["AI risk/anomaly layer", "Decision support", "Access-risk scoring, misuse anomaly detection, optional aggregate crime trend model; AI not final authority."],
        ["XAI layer", "Explanation", "Generates role-specific explanations; stores sensitive full explanations off-chain; writes explanation hash to audit."],
        ["Blockchain audit layer", "Tamper-evident audit", "Stores request IDs, policy/version hashes, decisions, approval proof hashes, model version, explanation hash, pointer hash, timestamps."],
        ["Encrypted off-chain storage", "Sensitive data protection", "Raw FIRs, witness/victim records, evidence media, full explanation text stay off-chain."],
        ["Audit/review dashboard", "Review", "Reconstructs decisions and detects tampered logs, missing explanations, stale credentials, abnormal requests, policy/version mismatch."],
        ["Data flow", "Step 1", "Officer submits subject, role/rank, station, jurisdiction, case ID, object, purpose, urgency."],
        ["Data flow", "Step 2", "Gateway validates credential and creates request ID."],
        ["Data flow", "Step 3", "Policy layer evaluates RBAC and ABAC/PBAC rules."],
        ["Data flow", "Step 4", "AI layer optionally scores access risk/anomaly likelihood."],
        ["Data flow", "Step 5", "XAI layer creates allow/deny/escalate explanation."],
        ["Data flow", "Step 6", "Superior officer reviews sensitive requests before final approval."],
        ["Data flow", "Step 7", "Blockchain audit layer records hashes and metadata."],
        ["Data flow", "Step 8", "If approved, gateway releases token or encrypted pointer."],
        ["Data flow", "Step 9", "Auditor reconstructs event and verifies hashes."],
    ]
    add_sheet(wb, "10_Architecture", ["Section", "Item", "Detail"], architecture_rows)

    methodology_rows = [
        ["Stage 1", "Synthetic multi-station workload", "Generate deterministic states, districts, stations, officers, cases, records, approvals, requests, audit events."],
        ["Stage 2", "Policy oracle", "Define ground-truth allow/deny/escalate decisions before any model metrics."],
        ["Stage 3", "Baselines", "RBAC + mutable log; ABAC/PBAC + mutable log; ABAC/PBAC + signed append-only log; basic Fabric-style audit."],
        ["Stage 4", "Proposed designs", "Fabric-style audit + ABAC/PBAC; add off-chain pointers; add superior approval; add XAI artifact logging."],
        ["Stage 5", "AI and XAI", "Use deterministic policies for final access; AI supports risk/anomaly scoring and explanation generation."],
        ["Stage 6", "Ablations", "Remove blockchain, ABAC/PBAC, off-chain encryption, superior approval, XAI logging, anomaly scoring; vary sensitive request ratios and outage conditions."],
        ["Stage 7", "India aggregate context", "Use NCRB/BPRD only for aggregate analysis, not individual prediction."],
        ["Stage 8", "Reproducibility", "Save config, seed, generator hash, policy hash, model hash, metrics, logs, audit events, explanations, plots, limitations."],
        ["Hypothesis H1", "Audit", "Fabric + ABAC/PBAC improves tamper detection and audit reconstruction versus centralized mutable logs."],
        ["Hypothesis H2", "Signed log baseline", "Signed append-only logs may be strong and faster than blockchain; report honestly if so."],
        ["Hypothesis H3", "XAI logging", "XAI artifact logging may improve reviewability and audit completeness but increase storage/latency."],
        ["Hypothesis H4", "Off-chain encryption", "Reduces raw data exposure but may not solve metadata leakage."],
        ["Hypothesis H5", "Public data", "NCRB/BPRD supports aggregate trend analysis, not individual-level prediction."],
    ]
    add_sheet(wb, "11_Methodology", ["Stage/Hypothesis", "Topic", "Detail"], methodology_rows)

    experiment_rows = [
        ["Experiment 1", "Access-control correctness", "Verify allow/deny/escalate decisions against policy oracle.", "authorization accuracy; false allow; false deny; false escalation; escalation precision; policy coverage; reason-code completeness"],
        ["Experiment 1", "Required scenarios", "Normal in-jurisdiction, cross-jurisdiction sensitive, revoked officer, stale case assignment, superior approval, juvenile/witness/victim sensitive, emergency, court/prosecutor, sealed record, expired approval.", ""],
        ["Experiment 2", "Audit completeness and tamper detection", "Test whether audit trail reconstructs events and detects manipulation.", "audit completeness; tamper detection; reconstruction success; missing events; hash verification failure; false tamper alert"],
        ["Experiment 2", "Tamper cases", "Delete log row, alter decision reason, alter explanation, alter pointer, replay approval, backdate request, revoke credential after request, remove approval, compromised station node.", ""],
        ["Experiment 3", "Latency, throughput, storage", "Measure cost of auditability under station/request/classified/cross-jurisdiction load settings.", "p50/p95/p99 latency; throughput; audit write latency; approval latency; storage/request; failed request rate"],
        ["Experiment 4", "Metadata leakage", "Measure exposed context even when raw records stay off-chain.", "metadata leakage score; re-identification proxy risk; sensitive-attribute inference; audit utility loss"],
        ["Experiment 5", "XAI reviewability", "Evaluate whether explanations are complete, stable, useful for audit.", "completeness; fidelity; stability; role coverage; hash verification; optional reviewer time"],
        ["Experiment 6", "Aggregate India crime trend context", "Use NCRB and BPRD for aggregate trend context only.", "MAE/RMSE; Poisson deviance; temporal holdout; per-state error; feature/explanation stability"],
        ["Experiment 7", "Optional security/blockchain benchmarks", "UNSW-NB15, CSE-CIC-IDS2018, Elliptic, Amazon Employee Access if time permits.", "security/anomaly/XAI generalization metrics"],
        ["Paper gate", "Before writing results", "Run one baseline and one proposed method; run ablations; save metrics; document failures; write limitations.", ""],
    ]
    add_sheet(wb, "12_Experiments", ["Experiment", "Topic", "Detail", "Metrics/Checks"], experiment_rows)

    metric_rows = [
        ["Access-control", "Authorization accuracy", "Fraction of requests matching the policy oracle."],
        ["Access-control", "False allow rate", "Denied/escalated requests incorrectly allowed; highest-risk error."],
        ["Access-control", "False deny rate", "Legitimate requests incorrectly denied."],
        ["Access-control", "False escalation rate", "Requests unnecessarily sent to superior review."],
        ["Access-control", "Escalation precision", "Fraction of escalations that truly required superior approval."],
        ["Access-control", "Policy coverage", "Fraction of requests with deterministic policy reason."],
        ["Access-control", "Reason-code completeness", "Fraction of decisions with machine-readable reason codes."],
        ["Security/Audit", "Audit completeness", "Required events present for request reconstruction."],
        ["Security/Audit", "Audit reconstruction success", "Decision reconstructable from stored request, policy, model, explanation, approval artifacts."],
        ["Security/Audit", "Tamper detection rate", "Injected tampering cases detected."],
        ["Security/Audit", "False tamper alert rate", "Benign cases incorrectly flagged."],
        ["Security/Audit", "Hash verification success", "Explanation/payload/policy hashes verify."],
        ["Security/Audit", "Revocation delay", "Time between credential revocation and denied access."],
        ["Blockchain/System", "Decision latency p50/p95/p99", "Request submit to allow/deny/escalate decision."],
        ["Blockchain/System", "Audit write latency p50/p95/p99", "Decision to committed audit event."],
        ["Blockchain/System", "Throughput", "Completed requests per second."],
        ["Blockchain/System", "Storage overhead per request", "Ledger bytes plus off-chain artifacts."],
        ["Privacy/Metadata", "Metadata leakage score", "Weighted exposure of station pair, role/rank, sensitivity, case type, timing, approval pattern."],
        ["Privacy/Metadata", "Sensitive-attribute inference accuracy", "Attacker's ability to infer hidden sensitivity/case type from audit metadata."],
        ["XAI", "Explanation completeness", "Required fields present: decision, decisive attributes, missing attributes, policy/model version, score if applicable."],
        ["XAI", "Explanation fidelity", "Agreement between explanation and actual model/policy logic."],
        ["XAI", "Explanation stability", "Consistency under small non-sensitive perturbations."],
        ["XAI", "Role coverage", "Officer, superior, auditor, court/prosecutor views available."],
        ["XAI", "Explanation verification", "Stored explanation hash matches artifact."],
        ["Aggregate crime", "MAE/RMSE", "Count or rate prediction error."],
        ["Aggregate crime", "Temporal holdout error", "Performance on later years held out from training."],
        ["Fairness/bias", "Disparate escalation rates", "Synthetic escalation differences across simulated station/district groups."],
    ]
    add_sheet(wb, "13_Metrics", ["Category", "Metric", "Definition"], metric_rows)

    ethics_rows = [
        ["Ethical position", "Accountability and safety", "Research must improve reviewable access, not expand surveillance or automate coercive policing."],
        ["High-risk data", "Victim/witness identities", "Highly sensitive."],
        ["High-risk data", "Juvenile records", "Highly sensitive."],
        ["High-risk data", "Sexual-offence and crimes against children records", "Highly sensitive."],
        ["High-risk data", "Caste/tribe related offences", "Highly sensitive."],
        ["High-risk data", "Trafficking, domestic violence, crimes against women", "Highly sensitive."],
        ["High-risk data", "Cybercrime victim reports, forensic reports, biometrics/device IDs", "Highly sensitive."],
        ["Security risk", "Insider misuse", "Mitigate with case assignment, purpose checks, superior approval, anomaly detection, audit review."],
        ["Security risk", "Metadata leakage", "Minimize on-chain metadata, hash IDs, test inference risk, selective visibility if needed."],
        ["Security risk", "False input problem", "Blockchain preserves bad input; require chain of custody, policy versioning, human review, correction process."],
        ["Security risk", "Explanation leakage", "Role-specific explanations, redact sensitive features, store full explanation off-chain, hash artifacts on-chain."],
        ["Privacy risk", "Linkability/re-identification/function creep", "Use minimal on-chain data and explicit privacy-risk testing."],
        ["Fairness risk", "Reported-crime bias and feedback loops", "Do not build individual risk scores; report aggregate limitations and caveats."],
        ["Human oversight", "Named decision-maker required", "Log requester, request, purpose, policy decision, AI score, explanation, approval, disclosure, expiry/revocation."],
        ["Allowed claim", "Design informed by Indian digital-policing and electronic-record context", "Allowed if cited."],
        ["Allowed claim", "Design avoids raw personal-data storage on-chain", "Allowed as architecture claim."],
        ["Not allowed", "System is DPDP compliant", "Do not claim without legal expert review."],
        ["Not allowed", "System guarantees evidence admissibility", "Do not claim."],
        ["Not allowed", "System deployable by Indian police", "Do not claim."],
        ["Not allowed", "System prevents misuse or is privacy preserving", "Do not claim without formal evidence."],
    ]
    add_sheet(wb, "14_Ethics_Legal", ["Category", "Item", "Detail"], ethics_rows)

    add_sheet(wb, "15_References", ["Group", "Label", "Original text", "Link"], extract_source_links())

    next_steps = [
        [1, "Inspect LEChain code", "Record reusable smart-contract/access-control ideas.", "Pending"],
        [2, "Locate BAXDT repository", "Search author pages, DOI supplementary links, and GitHub if available.", "Pending"],
        [3, "Implement DecisionTrace object", "BAXDT-style request/model/explanation/trace hash object.", "Pending"],
        [4, "Implement local hash-chain ledger", "Before full Hyperledger Fabric, build deterministic audit log baseline.", "Pending"],
        [5, "Generate synthetic access requests", "Start with 1,000 requests, seed 42.", "Pending"],
        [6, "Run baselines", "RBAC, ABAC/PBAC, signed-log, and SEBA-XAI local ledger variants.", "Pending"],
        [7, "Save artifacts", "Use experiments/runs, results/tables, results/plots, reports/iteration.", "Pending"],
        [8, "Write claims only after results", "No fabricated accuracy/security/performance claims.", "Required"],
    ]
    add_sheet(wb, "16_Next_Steps", ["Step", "Task", "Detail", "Status"], next_steps)

    source_rows = []
    for rel in SOURCE_FILES:
        path = ROOT / rel
        source_rows.append([rel, "exists" if path.exists() else "missing", str(path), path.stat().st_size if path.exists() else ""])
    add_sheet(wb, "17_Source_Files", ["Source file", "Status", "Absolute path", "Size bytes"], source_rows)

    add_sheet(wb, "18_Raw_MD_Lines", ["Source file", "Line no.", "Heading context", "Line text"], extract_markdown_lines(SOURCE_FILES))

    # Make important cells visually distinct.
    for ws_name in ["00_Read_Me", "01_Update_Map"]:
        ws = wb[ws_name]
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, str) and ("No implementation" in cell.value or "not completed" in cell.value or "No " in cell.value[:5]):
                    cell.fill = WARNING_FILL
                elif isinstance(cell.value, str) and ("Completed" in cell.value or "Selected" in cell.value):
                    cell.fill = OK_FILL

    wb.save(OUT_FILE)
    print(OUT_FILE)


if __name__ == "__main__":
    main()
