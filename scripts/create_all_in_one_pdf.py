from __future__ import annotations

import csv
import html
import textwrap
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
    XPreformatted,
)


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "pdf_exports"
OUT_FILE = OUT_DIR / "SEBA_XAI_All_In_One_Research_Pack.pdf"


MAIN_MARKDOWN = [
    "00_START_HERE.md",
    "README.md",
    "professor_update_2026-05-27.md",
    "research_brief.md",
    "00_problem_understanding.md",
    "01_literature_review.md",
    "03_datasets.md",
    "05_research_gap.md",
    "06_proposed_architecture.md",
    "07_methodology.md",
    "08_experiment_plan.md",
    "09_evaluation_metrics.md",
    "10_ethics_security_legal.md",
    "11_paper_outline.md",
    "12_five_day_learning_plan.md",
    "13_implementation_kickstart.md",
    "14_existing_models_for_testing.md",
]

SUPPORTING_MARKDOWN = [
    "sources/source_log.md",
    "sources/literature_matrix.md",
    "sources/dataset_inventory.md",
    "experiments/experiment_plan.md",
    "experiments/runs/README.md",
    "results/plots/README.md",
    "papers/final_paper/README.md",
    "papers/final_paper/introduction/README.md",
    "papers/final_paper/introduction/introduction_control_document.md",
    "papers/final_paper/introduction/introduction_skeleton.md",
    "papers/final_paper/introduction/introduction_draft_v0.md",
    "papers/final_paper/introduction/20_day_introduction_plan.md",
    "papers/final_paper/introduction/reviewer_objection_checklist.md",
]

COURSE_MARKDOWN = [
    "Learn/blockchain_xai_course/README.md",
    "Learn/blockchain_xai_course/SYLLABUS.md",
    "Learn/blockchain_xai_course/blockchain/00_overview.md",
    "Learn/blockchain_xai_course/blockchain/01_foundations.md",
    "Learn/blockchain_xai_course/blockchain/02_distributed_systems.md",
    "Learn/blockchain_xai_course/blockchain/03_bitcoin.md",
    "Learn/blockchain_xai_course/blockchain/04_ethereum.md",
    "Learn/blockchain_xai_course/blockchain/05_consensus.md",
    "Learn/blockchain_xai_course/blockchain/06_scalability_layer2.md",
    "Learn/blockchain_xai_course/blockchain/07_privacy_zk.md",
    "Learn/blockchain_xai_course/blockchain/08_defi.md",
    "Learn/blockchain_xai_course/blockchain/09_security_attacks.md",
    "Learn/blockchain_xai_course/blockchain/10_governance_tokenomics.md",
    "Learn/blockchain_xai_course/blockchain/11_advanced_consensus.md",
    "Learn/blockchain_xai_course/blockchain/12_research_frontiers.md",
    "Learn/blockchain_xai_course/xai/00_overview.md",
    "Learn/blockchain_xai_course/xai/01_foundations.md",
    "Learn/blockchain_xai_course/xai/02_taxonomy.md",
    "Learn/blockchain_xai_course/xai/03_feature_attribution.md",
    "Learn/blockchain_xai_course/xai/04_gradient_methods.md",
    "Learn/blockchain_xai_course/xai/05_concept_based.md",
    "Learn/blockchain_xai_course/xai/06_counterfactuals.md",
    "Learn/blockchain_xai_course/xai/07_global_methods.md",
    "Learn/blockchain_xai_course/xai/08_attention_transformers.md",
    "Learn/blockchain_xai_course/xai/09_mechanistic_interpretability.md",
    "Learn/blockchain_xai_course/xai/10_evaluation_metrics.md",
    "Learn/blockchain_xai_course/xai/11_human_centered.md",
    "Learn/blockchain_xai_course/xai/12_applications.md",
    "Learn/blockchain_xai_course/xai/13_research_frontiers.md",
    "Learn/blockchain_xai_course/intersection/blockchain_xai_intersection.md",
    "Learn/blockchain_xai_course/labs/README.md",
    "Learn/blockchain_xai_course/labs/lab_1_toy_blockchain.md",
    "Learn/blockchain_xai_course/labs/lab_2_reentrancy.md",
    "Learn/blockchain_xai_course/labs/lab_3_zk_circuit.md",
    "Learn/blockchain_xai_course/labs/lab_4_shap_from_scratch.md",
    "Learn/blockchain_xai_course/labs/lab_5_saliency_sanity.md",
    "Learn/blockchain_xai_course/labs/lab_6_zkml_prototype.md",
    "Learn/blockchain_xai_course/resources/books.md",
    "Learn/blockchain_xai_course/resources/courses.md",
    "Learn/blockchain_xai_course/resources/papers.md",
    "Learn/blockchain_xai_course/resources/tools.md",
]

REPORT_MARKDOWN = sorted(
    p.relative_to(ROOT).as_posix()
    for p in (ROOT / "reports" / "iteration").glob("*.md")
)

CSV_FILES = [
    "02_literature_matrix.csv",
    "04_dataset_matrix.csv",
    "results/tables/literature_matrix.csv",
    "results/tables/dataset_shortlist.csv",
]

SPREADSHEETS = [
    "spreadsheets/SEBA_XAI_Focused_Research_Only.xlsx",
    "spreadsheets/SEBA_XAI_Professor_Update_Details.xlsx",
]

EXCLUDE_DIRS_FOR_MD = {
    "professor_ready_documents",
    "research_documents",
}


styles = getSampleStyleSheet()
styles.add(
    ParagraphStyle(
        name="TitleCenter",
        parent=styles["Title"],
        alignment=TA_CENTER,
        fontSize=22,
        leading=28,
        spaceAfter=16,
    )
)
styles.add(
    ParagraphStyle(
        name="SectionHeading",
        parent=styles["Heading1"],
        fontSize=16,
        leading=20,
        spaceBefore=18,
        spaceAfter=8,
        textColor=colors.HexColor("#1F4E78"),
    )
)
styles.add(
    ParagraphStyle(
        name="SubHeading",
        parent=styles["Heading2"],
        fontSize=13,
        leading=16,
        spaceBefore=12,
        spaceAfter=6,
        textColor=colors.HexColor("#2F5597"),
    )
)
styles.add(
    ParagraphStyle(
        name="Small",
        parent=styles["BodyText"],
        fontSize=8,
        leading=10,
        spaceAfter=3,
    )
)
styles.add(
    ParagraphStyle(
        name="CodeSmall",
        parent=styles["Code"],
        fontSize=6.5,
        leading=8,
        leftIndent=8,
        spaceBefore=4,
        spaceAfter=4,
    )
)


def esc(text: object) -> str:
    return html.escape("" if text is None else str(text))


def para(text: str, style: str = "BodyText") -> Paragraph:
    return Paragraph(esc(text), styles[style])


def add_wrapped_code(story: list, text: str) -> None:
    wrapped_lines = []
    for line in text.splitlines():
        if not line:
            wrapped_lines.append("")
        else:
            wrapped_lines.extend(textwrap.wrap(line, width=110, replace_whitespace=False, drop_whitespace=False) or [""])
    story.append(XPreformatted(esc("\n".join(wrapped_lines)), styles["CodeSmall"]))


def add_markdown_file(story: list, rel: str, title_prefix: str = "") -> None:
    path = ROOT / rel
    if not path.exists():
        story.append(para(f"Missing file: {rel}", "BodyText"))
        return

    title = title_prefix + rel
    story.append(PageBreak())
    story.append(Paragraph(esc(title), styles["SectionHeading"]))
    story.append(para(f"Source file: {rel}", "Small"))

    text = path.read_text(encoding="utf-8")
    in_code = False
    code_lines: list[str] = []

    for raw in text.splitlines():
        line = raw.rstrip()
        if line.strip().startswith("```"):
            if in_code:
                add_wrapped_code(story, "\n".join(code_lines))
                code_lines = []
                in_code = False
            else:
                in_code = True
                code_lines = []
            continue

        if in_code:
            code_lines.append(line)
            continue

        stripped = line.strip()
        if not stripped:
            story.append(Spacer(1, 4))
        elif stripped.startswith("# "):
            story.append(Paragraph(esc(stripped[2:]), styles["SectionHeading"]))
        elif stripped.startswith("## "):
            story.append(Paragraph(esc(stripped[3:]), styles["SubHeading"]))
        elif stripped.startswith("### "):
            story.append(Paragraph(esc(stripped[4:]), styles["Heading3"]))
        elif stripped.startswith("|") and stripped.endswith("|"):
            add_wrapped_code(story, stripped)
        elif stripped.startswith("- "):
            story.append(Paragraph("• " + esc(stripped[2:]), styles["BodyText"]))
        elif stripped[:3].isdigit() and ". " in stripped[:5]:
            story.append(para(stripped))
        else:
            story.append(para(stripped))

    if in_code and code_lines:
        add_wrapped_code(story, "\n".join(code_lines))


def add_csv_file(story: list, rel: str) -> None:
    path = ROOT / rel
    if not path.exists():
        return
    story.append(PageBreak())
    story.append(Paragraph(esc(f"CSV Matrix: {rel}"), styles["SectionHeading"]))
    with path.open(newline="", encoding="utf-8") as fp:
        rows = list(csv.DictReader(fp))
    for idx, row in enumerate(rows, start=1):
        story.append(Paragraph(esc(f"Row {idx}"), styles["SubHeading"]))
        data = [[Paragraph(esc(k), styles["Small"]), Paragraph(esc(v), styles["Small"])] for k, v in row.items()]
        table = Table(data, colWidths=[4.2 * cm, 12.5 * cm])
        table.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#D9EAF7")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        story.append(table)
        story.append(Spacer(1, 6))


def add_spreadsheet_summary(story: list, rel: str) -> None:
    path = ROOT / rel
    if not path.exists():
        return
    story.append(PageBreak())
    story.append(Paragraph(esc(f"Spreadsheet Summary: {rel}"), styles["SectionHeading"]))
    wb = load_workbook(path, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        if sheet_name == "18_Raw_MD_Lines":
            story.append(Paragraph(esc(f"{sheet_name}: skipped in PDF because the source Markdown files are included directly."), styles["Small"]))
            continue
        ws = wb[sheet_name]
        story.append(Paragraph(esc(sheet_name), styles["SubHeading"]))
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        max_rows = min(len(rows), 35)
        for row in rows[:max_rows]:
            values = ["" if v is None else str(v) for v in row]
            add_wrapped_code(story, " | ".join(values))
        if len(rows) > max_rows:
            story.append(para(f"Only first {max_rows} rows shown here. Full sheet is available in the Excel file.", "Small"))


def add_inventory(story: list) -> None:
    story.append(PageBreak())
    story.append(Paragraph("Repository File Inventory", styles["SectionHeading"]))
    story.append(para("This inventory lists important files kept in the folder. Existing paper PDFs are listed, not embedded, to avoid duplicating source articles and generated PDF exports."))

    files = sorted(
        p.relative_to(ROOT).as_posix()
        for p in ROOT.rglob("*")
        if p.is_file() and ".git/" not in p.as_posix()
    )
    for rel in files:
        if rel.startswith("pdf_exports/"):
            continue
        story.append(para(rel, "Small"))


def footer(canvas, doc) -> None:
    canvas.saveState()
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.grey)
    canvas.drawString(1.4 * cm, 0.9 * cm, "SEBA-XAI All-in-One Research Pack")
    canvas.drawRightString(A4[0] - 1.4 * cm, 0.9 * cm, f"Page {doc.page}")
    canvas.restoreState()


def build_pdf() -> None:
    OUT_DIR.mkdir(exist_ok=True)
    doc = SimpleDocTemplate(
        str(OUT_FILE),
        pagesize=A4,
        rightMargin=1.35 * cm,
        leftMargin=1.35 * cm,
        topMargin=1.3 * cm,
        bottomMargin=1.3 * cm,
    )
    story: list = []

    story.append(Paragraph("SEBA-XAI All-in-One Research Pack", styles["TitleCenter"]))
    story.append(Paragraph("Secure Explainable Blockchain-Audited Access Overlay for Indian Police Data Sharing", styles["TitleCenter"]))
    story.append(Spacer(1, 16))
    story.append(para(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}"))
    story.append(para("Workspace: /Users/venkatrayudu/Workspace/Research/codex research"))
    story.append(Spacer(1, 12))
    story.append(Paragraph("Coverage Note", styles["SectionHeading"]))
    story.append(para("This PDF consolidates the source Markdown notes, CSV matrices, spreadsheet summaries, course notes, reports, and file inventory for the SEBA-XAI research folder."))
    story.append(para("Generated duplicate folders such as professor_ready_documents/ and research_documents/ are not repeated in full because their source notes are included directly. Existing external paper PDFs are listed in the inventory rather than embedded as full articles."))
    story.append(para("No experimental result, performance number, legal-compliance claim, deployment claim, or SOTA claim is added by this PDF."))

    story.append(PageBreak())
    story.append(Paragraph("Table of Contents", styles["SectionHeading"]))
    toc_sections = [
        "Main research notes",
        "Supporting evidence and final-paper workspace notes",
        "Blockchain + XAI course notes",
        "Iteration reports",
        "CSV matrices",
        "Spreadsheet summaries",
        "Repository file inventory",
    ]
    for item in toc_sections:
        story.append(para("• " + item))

    story.append(PageBreak())
    story.append(Paragraph("Main Research Notes", styles["SectionHeading"]))
    for rel in MAIN_MARKDOWN:
        add_markdown_file(story, rel)

    story.append(PageBreak())
    story.append(Paragraph("Supporting Evidence and Paper Workspace", styles["SectionHeading"]))
    for rel in SUPPORTING_MARKDOWN:
        add_markdown_file(story, rel)

    story.append(PageBreak())
    story.append(Paragraph("Blockchain + XAI Course Notes", styles["SectionHeading"]))
    for rel in COURSE_MARKDOWN:
        add_markdown_file(story, rel)

    story.append(PageBreak())
    story.append(Paragraph("Iteration Reports", styles["SectionHeading"]))
    for rel in REPORT_MARKDOWN:
        add_markdown_file(story, rel)

    story.append(PageBreak())
    story.append(Paragraph("CSV Matrices", styles["SectionHeading"]))
    for rel in CSV_FILES:
        add_csv_file(story, rel)

    story.append(PageBreak())
    story.append(Paragraph("Spreadsheet Summaries", styles["SectionHeading"]))
    for rel in SPREADSHEETS:
        add_spreadsheet_summary(story, rel)

    add_inventory(story)

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    print(OUT_FILE)


if __name__ == "__main__":
    build_pdf()
