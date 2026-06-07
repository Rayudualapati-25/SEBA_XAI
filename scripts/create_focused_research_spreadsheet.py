from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "spreadsheets"
OUT_FILE = OUT_DIR / "SEBA_XAI_Focused_Research_Only.xlsx"


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")
HEADER_FONT = Font(color="FFFFFF", bold=True)
LINK_FONT = Font(color="0563C1", underline="single")
BODY_FONT = Font(name="Calibri", size=11)
THIN = Side(style="thin", color="D9E2F3")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


LITERATURE_LINKS = {
    "CCTNS Operational Police Stations": "https://www.pib.gov.in/PressReleasePage.aspx?PRID=2238241",
    "Inter-Operable Criminal Justice System (ICJS)": "https://www.mha.gov.in/en/commoncontent/inter-operable-criminal-justice-system-icjs",
    "NIST SP 800-162: Guide to Attribute Based Access Control": "https://doi.org/10.6028/NIST.SP.800-162",
    "Hyperledger Fabric: A Distributed Operating System for Permissioned Blockchains": "https://doi.org/10.1145/3190508.3190538",
    "Attribute-based access control scheme for data sharing on Hyperledger Fabric": "https://doi.org/10.1016/j.jisa.2022.103182",
    "Two-Level Blockchain System for Digital Crime Evidence Management": "https://www.mdpi.com/1424-8220/21/9/3051",
    "LEChain: A blockchain-based lawful evidence management scheme for digital forensics": "https://doi.org/10.1016/j.future.2020.09.038",
    "Design and Implementation of a Digital Evidence Management Model Based on Hyperledger Fabric": "https://doi.org/10.3745/JIPS.04.0178",
    "Blockchain based access control systems: State of the art and challenges": "https://doi.org/10.1145/3350546.3352561",
    "Blockchain for Access Control Systems": "https://doi.org/10.6028/NIST.IR.8403",
    "Towards accountable and privacy-preserving blockchain-based access control for data sharing": "https://doi.org/10.1016/j.jisa.2024.103866",
    "Blockchain Based Auditable Access Control for Distributed Business Processes": "https://doi.org/10.1109/ICDCS47774.2020.00015",
    "Privacy-Preserving Machine Learning: Methods Challenges and Directions": "https://arxiv.org/abs/2108.04417",
    "Deep Learning with Differential Privacy": "https://arxiv.org/abs/1607.00133",
    "Stop explaining black box machine learning models for high stakes decisions and use interpretable models instead": "https://doi.org/10.1038/s42256-019-0048-x",
    "The accuracy fairness and limits of predicting recidivism": "https://doi.org/10.1126/sciadv.aao5580",
    "Fair Prediction with Disparate Impact": "https://doi.org/10.1089/big.2016.0047",
    "Inherent Trade-Offs in the Fair Determination of Risk Scores": "https://arxiv.org/abs/1609.05807",
    "Runaway Feedback Loops in Predictive Policing": "https://proceedings.mlr.press/v81/ensign18a.html",
    "Fundamental considerations for the use of explainable AI in law enforcement": "https://doi.org/10.3389/fpos.2025.1605619",
    "Dialogue-based XAI for Predictive Policing: a Field Study": "https://ceur-ws.org/Vol-4017/paper_03.pdf",
    "Blockchain-based auditing of legal decisions supported by explainable AI and generative AI tools": "https://doi.org/10.1016/j.engappai.2023.107666",
    "A Secure and Privacy-Preserving Blockchain-Based XAI-Justice System": "https://doi.org/10.3390/info14090477",
    "Machine learning in crime prediction": "",
    "Analysis of criminal spatial events in India using exploratory data analysis and regression": "https://doi.org/10.1016/j.compenvurbsys.2023.108761",
    "Forecasting Murder Motives in India Using Statistical Analysis and Explainable Artificial Intelligence": "https://doi.org/10.1109/UPCON62832.2024.10983398",
    "CriX: Intersection of Crime Demographics and Explainable AI": "",
    "IndianBailJudgments-1200: A Multi-Attribute Dataset for Legal NLP on Indian Bail Orders": "https://arxiv.org/abs/2507.02506",
}


def style_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
            if isinstance(cell.value, str) and cell.value.startswith("http"):
                cell.hyperlink = cell.value
                cell.font = LINK_FONT
            if isinstance(cell.value, str) and ("not claim" in cell.value.lower() or "reject" in cell.value.lower()):
                cell.fill = WARNING_FILL
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 90))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, min(max_len + 2, 60))


def add_sheet(wb: Workbook, title: str, headers: list[str], rows: list[list[object]]) -> None:
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_sheet(ws)


def read_csv_dicts(rel: str) -> list[dict[str, str]]:
    with (ROOT / rel).open(newline="", encoding="utf-8") as fp:
        return list(csv.DictReader(fp))


def main() -> None:
    OUT_DIR.mkdir(exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    problem_rows = [
        ["Core problem", "Indian police and criminal-justice agencies need to share sensitive records across police stations, districts, states, courts, prisons, forensic labs, and prosecution units.", "This creates a need for trusted, auditable, privacy-aware, and explainable access.", "00_problem_understanding.md"],
        ["Sensitive records", "FIR details; witness statements; victim records; juvenile information; forensic reports; cybercrime evidence; case diary material; inter-agency intelligence.", "These records should not be freely exposed or stored raw on-chain.", "00_problem_understanding.md"],
        ["Existing India baseline", "India already has CCTNS and ICJS-style infrastructure.", "The research must complement existing systems, not replace them.", "PIB CCTNS / MHA ICJS"],
        ["CCTNS evidence", "PIB release says all 17,798 police stations were using CCTNS as of 2026-02-01.", "Supports baseline infrastructure assumption.", "https://www.pib.gov.in/PressReleasePage.aspx?PRID=2238241"],
        ["ICJS evidence", "MHA describes ICJS as integrating Police/CCTNS, Courts/e-Courts, Jails/e-Prisons, Forensics/e-Forensics, and Prosecution/e-Prosecution through a Data Sharing Matrix.", "Supports inter-agency framing.", "https://www.mha.gov.in/en/commoncontent/icjsncrb-administration"],
        ["Research interpretation", "Build an intelligent secure overlay on top of CCTNS/ICJS-style infrastructure.", "Overlay focuses on auditability, access governance, privacy, and explanation.", "00_problem_understanding.md"],
        ["Blockchain role", "Tamper-evident commitments, approvals, policy-version records, access events, explanation hashes, and audit reconstruction.", "Blockchain is for audit proofs, not raw police data storage.", "00_problem_understanding.md"],
        ["Security role", "RBAC, ABAC/PBAC, encryption, credential revocation, off-chain storage, misuse/anomaly monitoring.", "Security decides and controls access.", "00_problem_understanding.md"],
        ["XAI role", "Explain allow, deny, escalate, or superior-approval decisions.", "XAI makes access decisions reviewable.", "00_problem_understanding.md"],
        ["Main decision", "When Officer A from Station X requests Record R from Station Y, should access be allowed, denied, or escalated, and can the system reconstruct why later?", "This is stronger than generic crime prediction.", "00_problem_understanding.md"],
        ["In scope", "Inter-station/inter-agency access workflows; sensitive record simulation; RBAC/ABAC/PBAC; permissioned audit; off-chain pointers; XAI; aggregate public India data.", "Defines what the first paper may cover.", "00_problem_understanding.md"],
        ["Out of scope", "Replacing CCTNS/ICJS; real deployment; individual suspect prediction from NCRB; automatic sensitive disclosure; legal-compliance claims; blockchain-alone privacy/security claims.", "Avoids overclaiming.", "00_problem_understanding.md"],
        ["RQ1", "Does permissioned blockchain + ABAC/PBAC improve audit completeness and tamper detection compared with centralized logs and signed append-only logs?", "Audit/security question.", "00_problem_understanding.md"],
        ["RQ2", "What latency, throughput, storage, and operational overhead does the audit overlay introduce?", "System-performance question.", "00_problem_understanding.md"],
        ["RQ3", "Can XAI-backed access justifications improve reviewability compared with raw policy outputs or opaque model scores?", "XAI question.", "00_problem_understanding.md"],
        ["RQ4", "How much sensitive metadata is exposed by different designs?", "Privacy/metadata-leakage question.", "00_problem_understanding.md"],
        ["RQ5", "Which India public datasets can support aggregate analysis without unsupported station-level or individual-level claims?", "Dataset-suitability question.", "00_problem_understanding.md"],
    ]
    add_sheet(wb, "Problem_Understanding", ["Section", "Detail", "Meaning / Why It Matters", "Source"], problem_rows)

    lit_rows = []
    for row in read_csv_dicts("02_literature_matrix.csv"):
        lit_rows.append([
            row["title"],
            row["year"],
            row["domain"],
            row["method"],
            row["dataset"],
            row["contribution"],
            row["limitation"],
            row["relation_to_idea"],
            row["evidence_strength"],
            LITERATURE_LINKS.get(row["title"], ""),
        ])
    add_sheet(
        wb,
        "Literature_Review",
        ["Title / Source", "Year", "Domain", "Method", "Dataset", "Contribution", "Limitation", "Relation to SEBA-XAI", "Evidence Strength", "Paper / Source Link"],
        lit_rows,
    )

    model_rows = [
        [1, "Blockchain-assisted explainable decision traces (BAXDT)", 2025, "Blockchain + XAI decision traces", "SHAP explanations; explanation-density metric; blockchain anchoring; model/data context.", "Very high for XAI artifact hashing and decision trace design.", "Rebuild/adapt decision-trace idea locally.", "Not police-specific; repository URL not verified from accessible metadata.", "https://doi.org/10.1016/j.knosys.2025.114402", "https://www.sciencedirect.com/science/article/pii/S0950705125014418", ""],
        [2, "Blockchain-based auditing of legal decisions supported by XAI and generative AI tools", 2024, "Blockchain + XAI + legal AI", "Audit trail for AI-assisted legal decisions; XAI artifacts; legal decision support.", "Very high for legal AI audit pattern.", "Use as closest conceptual and evaluation precedent.", "Not police-record access governance; no public repo found.", "https://doi.org/10.1016/j.engappai.2023.107666", "https://livrepository.liverpool.ac.uk/3179806/", ""],
        [3, "LEChain", 2021, "Blockchain + lawful evidence management", "Consortium blockchain; evidence lifecycle; CP-ABE; witness/juror privacy.", "High for police/legal evidence and access-control workflow.", "Inspect and adapt evidence/access-control ideas.", "No XAI; focuses on evidence management rather than access explanation.", "https://doi.org/10.1016/j.future.2020.09.038", "https://www.sciencedirect.com/science/article/pii/S0167739X1933167X", "https://github.com/SopmmmodII/LEChain"],
        [4, "Two-Level Blockchain System for Digital Crime Evidence Management", 2021, "Digital crime evidence management", "Police/prosecutor/court consortium; Hyperledger Fabric; hot/cold blockchain separation.", "High for police/prosecutor/court blockchain architecture.", "Borrow hot/cold split and consortium design.", "No XAI layer; no public code found.", "https://doi.org/10.3390/s21093051", "https://www.mdpi.com/1424-8220/21/9/3051", ""],
        [5, "User authentication and access control to blockchain-based forensic log data", 2023, "Forensic log access control", "Authentication; access control; blockchain forensic logs; AVISPA-style verification.", "Medium for security baseline.", "Borrow RBAC/ABAC and formal verification ideas.", "No XAI; no public repo verified.", "https://doi.org/10.1186/s13635-023-00142-3", "https://link.springer.com/article/10.1186/s13635-023-00142-3", ""],
        [6, "BlendSPS", 2020, "Smart public safety blockchain", "Public-safety microservices; blockchain security service; Ethereum/Tendermint prototype.", "Medium for public-safety blockchain architecture.", "Use for modular architecture inspiration.", "Surveillance/IoT focus; not police records; no XAI.", "https://doi.org/10.3390/smartcities3030047", "https://www.mdpi.com/2624-6511/3/3/47", "https://github.com/samuelxu999/Research/tree/master/Security/py_dev/BlendSPS/"],
        [7, "A Secure and Privacy-Preserving Blockchain-Based XAI-Justice System", 2023, "Blockchain + privacy + XAI justice", "Privacy techniques; XAI; NLP/legal decision support.", "Medium for justice-domain framing.", "Cite as related vision.", "Conceptual; not an empirical baseline.", "https://doi.org/10.3390/info14090477", "https://www.mdpi.com/2078-2489/14/9/477", ""],
        [8, "Blockchain for explainable and trustworthy artificial intelligence", "2019/2020", "General blockchain + XAI", "Smart contracts; trusted oracles; decentralized storage; AI/XAI predictor consensus.", "Medium for general XAI+blockchain theory.", "Cite for general rationale only.", "Not police/legal access-governance implementation.", "https://doi.org/10.1002/widm.1340", "https://scholarworks.aub.edu.lb/handle/10938/25595?show=full", ""],
    ]
    add_sheet(
        wb,
        "Existing_Model_Search",
        ["Rank", "System / Paper", "Year", "Domain", "Components Present", "Fit for SEBA-XAI", "Recommended Use", "Limitation", "DOI", "Source Link", "Code Link"],
        model_rows,
    )

    dataset_rows = []
    for row in read_csv_dicts("04_dataset_matrix.csv"):
        dataset_rows.append([
            row["dataset"],
            row["source"],
            row["fields_or_content"],
            row["granularity"],
            row["usability"],
            row["limitations"],
            row["publication_suitability"],
            row["recommended_use"],
        ])
    add_sheet(
        wb,
        "Dataset_Study",
        ["Dataset", "Source Link", "Fields / Content", "Granularity", "Usability", "Limitations", "Publication Suitability", "Recommended Use"],
        dataset_rows,
    )

    gap_rows = [
        ["Already known", "India already has CCTNS/ICJS digital policing and justice infrastructure.", "Do not claim to create the first India police data-sharing system.", "PIB CCTNS / MHA ICJS"],
        ["Already known", "Blockchain has already been proposed for digital evidence and chain of custody.", "Need a more specific contribution than simply using blockchain.", "Two-Level Blockchain / LEChain"],
        ["Already known", "Hyperledger Fabric, ABAC, blockchain access control, XAI/fairness, and NCRB crime analytics already exist.", "The novelty must combine them in a specific access-governance workflow.", "05_research_gap.md"],
        ["Reject", "Put all police data on blockchain.", "Rejected because raw FIRs, witness/victim records, juvenile records, and forensic records are too sensitive for raw on-chain storage.", "05_research_gap.md"],
        ["Reject", "AI crime prediction for India using public NCRB as the main contribution.", "Rejected because NCRB is aggregate reported/registered crime data, not individual suspect or incident-level police data.", "05_research_gap.md"],
        ["Reject", "CCTNS replacement.", "Rejected because CCTNS/ICJS already exist at national scale and replacement is unrealistic for this paper.", "05_research_gap.md"],
        ["Reject", "XAI solves trust.", "Rejected because XAI improves reviewability but can still be incomplete, misleading, or sensitive.", "05_research_gap.md"],
        ["Strong gap", "Limited reproducible work on a CCTNS/ICJS-compatible secure overlay that jointly evaluates blockchain audit, ABAC/PBAC, superior approval, off-chain encryption, XAI artifacts, tamper tests, and metadata leakage.", "This is the selected publishable gap.", "05_research_gap.md"],
        ["Problem statement", "How can an intelligent secure overlay support auditable, privacy-aware, and explainable access to sensitive police/criminal-justice records while preserving CCTNS/ICJS-style infrastructure and avoiding raw sensitive data on-chain?", "This is the current research problem.", "05_research_gap.md"],
        ["Proposed title", "SEBA-XAI: A Secure, Explainable, Blockchain-Audited Access Overlay for Inter-Agency Police Data Sharing in India.", "Working title.", "05_research_gap.md"],
        ["Objective 1", "Keep sensitive police data off-chain and store only audit proofs on the blockchain.", "Simple architecture goal.", "05_research_gap.md"],
        ["Objective 2", "Create sample police-station access requests for testing because real CCTNS/ICJS request logs are not public.", "Simple data/testing goal.", "05_research_gap.md"],
        ["Objective 3", "Compare normal access-control logs with the proposed blockchain-audited access system.", "Simple comparison goal.", "05_research_gap.md"],
        ["Objective 4", "Check whether the system makes correct allow/deny/escalate decisions and whether tampering can be detected.", "Simple evaluation goal.", "05_research_gap.md"],
        ["Objective 5", "Use NCRB and BPRD only for background and aggregate analysis, not for predicting individual criminals.", "Simple dataset boundary.", "05_research_gap.md"],
        ["Final recommendation", "Build the first paper as a secure systems plus responsible AI paper, not a pure ML crime-prediction paper.", "Best professor-facing framing.", "05_research_gap.md"],
    ]
    add_sheet(wb, "Research_Gap", ["Category", "Point", "Meaning", "Source"], gap_rows)

    rq_terms = [
        ["Permissioned blockchain", "A blockchain where only approved organizations or users can participate.", "In this research, police stations or agencies would be known members, not anonymous public users."],
        ["ABAC", "Attribute-Based Access Control. Access is decided using attributes such as role, station, case assignment, record sensitivity, purpose, and time.", "Useful because police access depends on more than just job title."],
        ["PBAC", "Policy-Based Access Control. Access is decided using written rules or policies.", "Useful for rules such as 'juvenile records need superior approval'."],
        ["Overlay", "An extra layer added on top of an existing system.", "SEBA-XAI is an overlay because it supports CCTNS/ICJS-style systems instead of replacing them."],
        ["Audit completeness", "How much of the required access history is available for review.", "If a request is reviewed later, the system should show who requested, why, what decision was made, and what approval was used."],
        ["Tamper detection", "The ability to notice if someone changed, deleted, or replayed a log, approval, explanation, or record pointer.", "This is where blockchain or hash-chain logging becomes useful."],
        ["Centralized log", "A normal log stored in one database or system.", "Easy to implement, but weaker if an insider or compromised admin can edit it."],
        ["Signed append-only log", "A log where each new entry is linked using signatures or hashes and old entries should not be edited.", "This is a strong non-blockchain baseline."],
        ["Fabric-only audit", "A Hyperledger Fabric-style ledger that records audit events without all extra privacy and XAI features.", "Useful as a blockchain baseline."],
        ["Off-chain storage", "Keeping raw sensitive data outside the blockchain.", "Raw FIRs, victim details, witness details, and forensic records should stay off-chain."],
        ["Encrypted off-chain storage", "Off-chain data is protected with encryption.", "Only authorized users should get access to the decrypted data."],
        ["Metadata", "Information about data, such as who requested it, when, from which station, record type, sensitivity level, and approval status.", "Even metadata can reveal sensitive investigation patterns."],
        ["Metadata leakage", "Sensitive information indirectly exposed through metadata even when raw records are hidden.", "Example: repeated requests between two stations may reveal an active investigation."],
        ["XAI", "Explainable AI. It explains why a model or decision system gave a result.", "In this research, XAI explains allow, deny, or escalate decisions."],
        ["XAI-backed access justification", "An explanation attached to an access decision.", "Example: 'Denied because credential is revoked' or 'Escalated because record has witness-sensitive flag'."],
        ["Reviewability", "How easily a human reviewer can understand and check a decision.", "Important for officers, superior officers, auditors, and legal reviewers."],
        ["Opaque model score", "A model output number without a clear reason.", "Example: risk score 0.82 without saying why."],
        ["Policy output", "The direct result of applying access rules.", "Example: allow, deny, or escalate."],
        ["Access-risk score", "A score showing how unusual or risky an access request looks.", "It should support review, not automatically disclose sensitive data."],
        ["Allow decision", "The request satisfies the policy and access can be granted.", "Example: correct officer, assigned case, correct jurisdiction, valid purpose."],
        ["Deny decision", "The request violates a rule and access should not be granted.", "Example: revoked credential or sealed record without authority."],
        ["Escalate decision", "The request may be valid but needs superior approval.", "Example: juvenile, witness-sensitive, victim-sensitive, or cross-jurisdiction classified record."],
        ["Synthetic workload", "Artificial test data created using documented rules.", "Needed because real CCTNS/ICJS access-request logs are not public."],
        ["Aggregate dataset", "Data grouped by region, year, crime type, or table instead of individual records.", "NCRB data can support broad analysis, not individual prediction."],
        ["Station-level claim", "A claim about a specific police station's operation or prediction.", "Should not be made unless station-level evidence is available."],
        ["Individual-level claim", "A claim about a specific person, suspect, victim, or officer.", "Should not be made from public aggregate NCRB data."],
        ["Operational overhead", "Extra time, storage, cost, or complexity added by the system.", "Blockchain audit and XAI explanations may improve review but add overhead."],
        ["Latency", "How long one request takes from submission to decision or audit write.", "Important for practical system use."],
        ["Throughput", "How many requests the system can handle per second or per time period.", "Important when many stations or agencies send requests."],
        ["Storage overhead", "Extra storage needed for logs, hashes, explanations, and audit artifacts.", "Important because audit and XAI artifacts can grow over time."],
        ["Revocation", "Removing or disabling a user's credential or approval.", "Important if an officer is no longer authorized."],
        ["Revocation delay", "Time taken for a revocation to affect access decisions.", "A long delay can allow unauthorized access."],
        ["Explanation stability", "Whether explanations remain consistent when small non-sensitive changes are made.", "Unstable explanations are harder to trust or audit."],
    ]
    add_sheet(wb, "Research_Question_Terms", ["Term", "Simple Definition", "Meaning in This Research"], rq_terms)

    wb.save(OUT_FILE)
    print(OUT_FILE)


if __name__ == "__main__":
    main()
