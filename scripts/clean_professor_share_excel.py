"""Remove least-relevant rows from professor-share Excel workbooks.

The cleanup keeps the professor package focused on the research direction:
CCTNS/ICJS-compatible access governance, permissioned blockchain audit,
ABAC/PBAC security, privacy, and XAI for reviewable access decisions.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SHARE = ROOT / "professor_share_research_only_2026-05-29"

DROP_LITERATURE_TITLES = {
    "Deep Learning with Differential Privacy",
    "Forecasting Murder Motives in India Using Statistical Analysis and Explainable Artificial Intelligence",
    "CriX: Intersection of Crime Demographics and Explainable AI",
    "IndianBailJudgments-1200: A Multi-Attribute Dataset for Legal NLP on Indian Bail Orders",
}

DROP_MODEL_NAMES = {
    "Blockchain for explainable and trustworthy artificial intelligence",
}

DROP_DATASETS = {
    "Synthetic multi-station access-control workload",
    "Synthetic multi-station ABAC workload",
    "NCRB ADSI 2023",
    "UK Police Open Data",
    "Chicago Crimes 2001 to Present",
    "FBI NIBRS / Crime Data API",
    "UCI Communities and Crime",
    "Elliptic Bitcoin Transaction Dataset",
    "UNSW-NB15",
    "CSE-CIC-IDS2018",
    "Amazon Employee Access Challenge",
    "Kaggle Indian crime mirrors",
}


def autosize(ws) -> None:
    for col in ws.columns:
        width = 12
        letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                width = max(width, min(len(str(cell.value)) + 2, 65))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions[letter].width = width
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    ws.freeze_panes = "A2"


def delete_rows_by_first_col(ws, drop_values: set[str]) -> int:
    removed = 0
    for row_idx in range(ws.max_row, 1, -1):
        value = ws.cell(row_idx, 1).value
        if value is not None and str(value).strip() in drop_values:
            ws.delete_rows(row_idx)
            removed += 1
    return removed


def delete_rows_by_second_col(ws, drop_values: set[str]) -> int:
    removed = 0
    for row_idx in range(ws.max_row, 1, -1):
        value = ws.cell(row_idx, 2).value
        if value is not None and str(value).strip() in drop_values:
            ws.delete_rows(row_idx)
            removed += 1
    return removed


def delete_blank_primary_rows(ws) -> int:
    removed = 0
    for row_idx in range(ws.max_row, 1, -1):
        value = ws.cell(row_idx, 1).value
        if value is None or not str(value).strip():
            ws.delete_rows(row_idx)
            removed += 1
    return removed


def clean_literature_matrix(path: Path) -> int:
    wb = load_workbook(path)
    removed = 0
    ws = wb["Literature_Matrix"]
    removed += delete_rows_by_first_col(ws, DROP_LITERATURE_TITLES)
    removed += delete_blank_primary_rows(ws)
    autosize(ws)
    wb.save(path)
    return removed


def clean_dataset_matrix(path: Path) -> int:
    wb = load_workbook(path)
    removed = 0
    if "Dataset_Matrix" in wb.sheetnames:
        removed += delete_rows_by_first_col(wb["Dataset_Matrix"], DROP_DATASETS)
        removed += delete_blank_primary_rows(wb["Dataset_Matrix"])
        autosize(wb["Dataset_Matrix"])
    if "Dataset_Shortlist" in wb.sheetnames:
        removed += delete_rows_by_second_col(wb["Dataset_Shortlist"], DROP_DATASETS)
        rebuild_dataset_shortlist(wb["Dataset_Shortlist"])
        autosize(wb["Dataset_Shortlist"])
    wb.save(path)
    return removed


def rebuild_dataset_shortlist(ws) -> None:
    """Replace the old broad shortlist with a focused professor-facing shortlist."""
    ws.delete_rows(1, ws.max_row)
    rows = [
        [
            "rank",
            "dataset",
            "source_url",
            "source_type",
            "content",
            "granularity",
            "access_license_note",
            "first_experiment_use",
            "limitations",
        ],
        [
            1,
            "NCRB Crime in India 2023",
            "https://www.data.gov.in/catalog/crime-india-2023",
            "official India OGD",
            "Aggregate crime tables by crime category, state/UT, city, arrests, disposal, and convictions where available.",
            "annual aggregate",
            "Public government dataset; check table-specific license notes before publication.",
            "India context and aggregate trend/background analysis.",
            "Reported/registered cases only; not suitable for individual prediction.",
        ],
        [
            2,
            "BPRD Data on Police Organizations",
            "https://bprd.nic.in/en/page/data_on_police_organization_dopo",
            "official India report",
            "Police manpower, infrastructure, stations, vehicles, and organization statistics.",
            "state/UT and organization level",
            "Public official report; cite BPRD source and year.",
            "Police resource and institutional context.",
            "Aggregate resource data, not access-request or FIR-level data.",
        ],
        [
            3,
            "CCTNS operational status",
            "https://www.pib.gov.in/PressReleasePage.aspx?PRID=2238241",
            "official India source",
            "CCTNS coverage, police-station digitization, FIR/chargesheet process context, replication, and standardized codes.",
            "national/system status",
            "Official press release; cite exact date and figure.",
            "Introduction baseline for existing digital policing infrastructure.",
            "Not a research dataset and not raw CCTNS data.",
        ],
        [
            4,
            "ICJS official description",
            "https://www.mha.gov.in/en/commoncontent/icjsncrb-administration",
            "official India source",
            "Police, courts, prisons, forensics, prosecution integration and Data Sharing Matrix description.",
            "system-level description",
            "Official MHA source.",
            "Introduction and architecture baseline for inter-agency sharing.",
            "No public operational data or implementation details.",
        ],
        [
            5,
            "I4C National Cybercrime Reporting Portal",
            "https://i4c.mha.gov.in/ncrp.aspx",
            "official India source",
            "Cybercrime complaint reporting, workflow, dashboards, tracking, and financial fraud reporting context.",
            "system-level description",
            "Official government portal.",
            "Cybercrime-sensitive-record motivation and workflow context.",
            "Not an open incident-level dataset.",
        ],
        [
            6,
            "NJDG/eCourts public judicial data",
            "https://doj.gov.in/the-national-judicial-data-grid-njdg/",
            "official India source",
            "Court case status, judicial data, orders, and related justice-system context where public.",
            "court/case/judgment level depending source",
            "Use only public court/judgment data with source-specific terms.",
            "Court/prosecution-side context for ICJS-style sharing.",
            "Not police access-request data.",
        ],
        [
            7,
            "IndianBailJudgments-1200",
            "https://arxiv.org/abs/2507.02506",
            "academic dataset",
            "Indian bail judgments with legal attributes such as bail outcome, crime type, IPC sections, and reasoning.",
            "judgment/case level",
            "Use only after checking dataset license and access terms.",
            "Optional legal NLP/XAI reference, not the main experiment.",
            "Judicial text, not police record access logs.",
        ],
        [
            8,
            "ILDC for CJPE",
            "https://arxiv.org/abs/2105.13562",
            "academic dataset",
            "Indian Supreme Court legal corpus for judgment prediction/explanation research.",
            "case/judgment level",
            "Use only under dataset license terms.",
            "Optional legal NLP/XAI background.",
            "Not specific to police access governance.",
        ],
    ]
    for row in rows:
        ws.append(row)


def clean_focused_workbook(path: Path) -> int:
    wb = load_workbook(path)
    removed = 0
    if "Literature_Review" in wb.sheetnames:
        removed += delete_rows_by_first_col(wb["Literature_Review"], DROP_LITERATURE_TITLES)
        removed += delete_blank_primary_rows(wb["Literature_Review"])
        autosize(wb["Literature_Review"])
    if "Existing_Model_Search" in wb.sheetnames:
        removed += delete_rows_by_second_col(wb["Existing_Model_Search"], DROP_MODEL_NAMES)
        autosize(wb["Existing_Model_Search"])
    if "Dataset_Study" in wb.sheetnames:
        removed += delete_rows_by_first_col(wb["Dataset_Study"], DROP_DATASETS)
        removed += delete_blank_primary_rows(wb["Dataset_Study"])
        autosize(wb["Dataset_Study"])
    for name in wb.sheetnames:
        if name not in {"Literature_Review", "Existing_Model_Search", "Dataset_Study"}:
            autosize(wb[name])
    wb.save(path)
    return removed


def main() -> None:
    removals: list[tuple[str, int]] = []
    removals.append(
        ("08_Literature_Matrix.xlsx", clean_literature_matrix(SHARE / "08_Literature_Matrix.xlsx"))
    )
    removals.append(("09_Dataset_Matrix.xlsx", clean_dataset_matrix(SHARE / "09_Dataset_Matrix.xlsx")))
    removals.append(
        ("12_Focused_Research_Only.xlsx", clean_focused_workbook(SHARE / "12_Focused_Research_Only.xlsx"))
    )
    for filename, count in removals:
        print(f"{filename}: removed {count} least-relevant rows")


if __name__ == "__main__":
    main()
